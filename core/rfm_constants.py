"""
تولید فایل rfm_constant.xlsx بر اساس داده‌های جدول rfm_data.

این فایل هم برای کاربر قابل‌فهم است (معیارها و لیبل‌ها)
و هم ساختار machine-readable دارد تا در مراحل بعد برای محاسبه RFM خوانده شود.
"""
from pathlib import Path

import jdatetime
from openpyxl import Workbook

from config import RFM_QUANTILE_BANDS
from core.db_manager import SQLiteManager
from core.rfm_data import RFM_DATA_TABLE


def _metric_labels(metric: str, score: int, max_score: int) -> str:
    recency_labels = {
        5: "خیلی جدید",
        4: "جدید",
        3: "متوسط",
        2: "قدیمی",
        1: "خیلی قدیمی",
    }
    frequency_labels = {
        5: "خیلی پرتکرار",
        4: "پرتکرار",
        3: "متوسط",
        2: "کم‌تکرار",
        1: "خیلی کم‌تکرار",
    }
    monetary_labels = {
        5: "خیلی ارزشمند",
        4: "ارزشمند",
        3: "متوسط",
        2: "کم‌ارزش",
        1: "خیلی کم‌ارزش",
    }
    mapping = {
        "recency_days": recency_labels,
        "total_orders": frequency_labels,
        "total_spent": monetary_labels,
    }
    label = mapping.get(metric, {}).get(score, "")
    if label:
        return label
    return f"سطح {score} از {max_score}"


def _metric_fa_name(metric: str) -> str:
    return {
        "recency_days": "تازگی خرید (روز)",
        "total_orders": "تعداد سفارش",
        "total_spent": "ارزش کل خرید",
    }.get(metric, metric)


def _fetch_metric_bands(
    db: SQLiteManager,
    metric: str,
    quantile_bands: int,
) -> list[tuple[int, float, float, int]]:
    """
    تقسیم داده‌های هر معیار به باندهای Quantile با NTILE.
    خروجی: [(bucket, min_val, max_val, cnt), ...]
    """
    sql = f"""
WITH ranked AS (
    SELECT
        CAST("{metric}" AS REAL) AS val,
        NTILE({quantile_bands}) OVER (ORDER BY CAST("{metric}" AS REAL) ASC) AS bucket
    FROM "{RFM_DATA_TABLE}"
    WHERE "{metric}" IS NOT NULL
)
SELECT
    bucket,
    MIN(val) AS min_val,
    MAX(val) AS max_val,
    COUNT(*) AS cnt
FROM ranked
GROUP BY bucket
ORDER BY bucket;
"""
    rows = db.execute(sql).fetchall()
    return [(int(r[0]), float(r[1]), float(r[2]), int(r[3])) for r in rows]


def _fetch_metric_stats(db: SQLiteManager, metric: str) -> tuple[int, float | None, float | None, float | None]:
    sql = f"""
SELECT
    COUNT("{metric}") AS cnt,
    MIN(CAST("{metric}" AS REAL)) AS min_v,
    MAX(CAST("{metric}" AS REAL)) AS max_v,
    AVG(CAST("{metric}" AS REAL)) AS avg_v
FROM "{RFM_DATA_TABLE}";
"""
    cnt, min_v, max_v, avg_v = db.execute(sql).fetchone()
    return int(cnt or 0), min_v, max_v, avg_v


def create_rfm_constant_excel(db: SQLiteManager, output_folder: str | Path) -> Path:
    """
    ساخت فایل rfm_constant.xlsx کنار خروجی rfm_data.
    """
    output_folder = Path(output_folder)
    output_folder.mkdir(parents=True, exist_ok=True)
    output_path = output_folder / "rfm_constant.xlsx"
    quantile_bands = max(2, int(RFM_QUANTILE_BANDS))

    total_rows = db.execute(f'SELECT COUNT(*) FROM "{RFM_DATA_TABLE}"').fetchone()[0]

    wb = Workbook()

    # Sheet 1: راهنما/متادیتا
    ws_meta = wb.active
    ws_meta.title = "meta"
    ws_meta.sheet_view.rightToLeft = True
    ws_meta.append(["کلید", "مقدار"])
    ws_meta.append(["table_name", RFM_DATA_TABLE])
    ws_meta.append(["generated_at_shamsi", jdatetime.datetime.now().strftime("%Y/%m/%d %H:%M:%S")])
    ws_meta.append(["total_rows", total_rows])
    ws_meta.append(["quantile_bands", quantile_bands])
    ws_meta.append(["note", f"باندها با NTILE({quantile_bands}) از داده‌های فعلی ساخته شده‌اند."])
    ws_meta.append(["note2", "در recency هرچه مقدار کمتر باشد، امتیاز بالاتر است."])

    # Sheet 2: آستانه‌ها (machine-readable)
    ws_thr = wb.create_sheet("thresholds")
    ws_thr.sheet_view.rightToLeft = True
    ws_thr.append(
        [
            "metric",
            "metric_fa",
            "bucket",
            "quantile_label",
            "percentile_from",
            "percentile_to",
            "score",
            "min_value",
            "max_value",
            "sample_count",
            "label_fa",
            "scoring_direction",
            "rule_text",
        ]
    )

    metrics = ["recency_days", "total_orders", "total_spent"]
    for metric in metrics:
        bands = _fetch_metric_bands(db, metric, quantile_bands)
        for bucket, min_val, max_val, cnt in bands:
            # recency: bucket پایین‌تر => score بالاتر
            if metric == "recency_days":
                score = quantile_bands + 1 - bucket
                scoring_direction = "lower_is_better"
            else:
                score = bucket
                scoring_direction = "higher_is_better"
            label = _metric_labels(metric, score, quantile_bands)
            q_label = f"Q{bucket}"
            percentile_from = round(((bucket - 1) / quantile_bands) * 100, 2)
            percentile_to = round((bucket / quantile_bands) * 100, 2)
            rule_text = f"{min_val:.2f} <= {metric} <= {max_val:.2f}"
            ws_thr.append(
                [
                    metric,
                    _metric_fa_name(metric),
                    bucket,
                    q_label,
                    percentile_from,
                    percentile_to,
                    score,
                    min_val,
                    max_val,
                    cnt,
                    label,
                    scoring_direction,
                    rule_text,
                ]
            )

    # Sheet 3: آمار کلی هر معیار (برای فهم کاربر)
    ws_stats = wb.create_sheet("metric_stats")
    ws_stats.sheet_view.rightToLeft = True
    ws_stats.append(["metric", "metric_fa", "count", "min", "max", "avg"])
    for metric in metrics:
        cnt, min_v, max_v, avg_v = _fetch_metric_stats(db, metric)
        ws_stats.append([metric, _metric_fa_name(metric), cnt, min_v, max_v, avg_v])

    # Sheet 4: لیبل/سگمنت پیشنهادی بر اساس score ranges
    ws_seg = wb.create_sheet("segment_rules")
    ws_seg.sheet_view.rightToLeft = True
    ws_seg.append(["segment", "r_min", "r_max", "f_min", "f_max", "m_min", "m_max", "description"])
    if quantile_bands == 5:
        ws_seg.append(["Champions", 4, 5, 4, 5, 4, 5, "خریداران بسیار ارزشمند و فعال"])
        ws_seg.append(["Loyal Customers", 3, 5, 4, 5, 3, 5, "مشتریان وفادار با خرید مستمر"])
        ws_seg.append(["Potential Loyalist", 4, 5, 2, 3, 2, 5, "جدید/روبه‌رشد، مناسب پرورش وفاداری"])
        ws_seg.append(["At Risk", 1, 2, 3, 5, 3, 5, "قبلاً خوب بوده‌اند اما اخیراً افت کرده‌اند"])
        ws_seg.append(["Hibernating", 1, 2, 1, 2, 1, 2, "غیرفعال یا کم‌ارزش، نیازمند کمپین فعال‌سازی"])
    else:
        high_min = max(quantile_bands - 1, 1)
        mid_min = max(quantile_bands // 2, 1)
        low_max = min(2, quantile_bands)
        ws_seg.append(
            [
                "Top Value",
                high_min,
                quantile_bands,
                high_min,
                quantile_bands,
                high_min,
                quantile_bands,
                "تعریف عمومی برای باندهای بالایی (غیر ۵-تایی).",
            ]
        )
        ws_seg.append(
            [
                "Mid Value",
                mid_min,
                quantile_bands,
                mid_min,
                quantile_bands,
                mid_min,
                quantile_bands,
                "تعریف عمومی برای کاربران میانی.",
            ]
        )
        ws_seg.append(
            [
                "Low Value",
                1,
                low_max,
                1,
                low_max,
                1,
                low_max,
                "تعریف عمومی برای کاربران کم‌ارزش.",
            ]
        )

    # اگر داده‌ای نبود، یک هشدار واضح داخل فایل بگذار
    if total_rows == 0:
        ws_meta.append(["warning", "جدول rfm_data خالی است؛ آستانه‌ها ممکن است کامل نباشند."])

    wb.save(str(output_path))
    return output_path
