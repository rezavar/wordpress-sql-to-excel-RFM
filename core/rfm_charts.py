"""
ساخت نمودارهای RFM از روی rfm_scores.xlsx و rfm_constant.xlsx.
خروجی در پوشه charts داخل پوشه خروجی.
"""
from pathlib import Path

import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import squarify
from openpyxl import load_workbook


def _to_int(value):
    if value is None:
        return None
    try:
        return int(float(str(value).strip()))
    except (ValueError, TypeError):
        return None


def _load_segment_rules(constant_path: Path) -> list[tuple[str, int, int, int, int, int, int]]:
    """
    خواندن شیت segment_rules از rfm_constant.xlsx.
    برمی‌گرداند: [(segment, r_min, r_max, f_min, f_max, m_min, m_max), ...]
    """
    wb = load_workbook(constant_path, read_only=True, data_only=True)
    if "segment_rules" not in wb.sheetnames:
        wb.close()
        return []
    ws = wb["segment_rules"]
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header:
        wb.close()
        return []
    headers = [str(c).strip() if c is not None else "" for c in header]
    idx = {h: i for i, h in enumerate(headers)}
    required = ["segment", "r_min", "r_max", "f_min", "f_max", "m_min", "m_max"]
    if not all(k in idx for k in required):
        wb.close()
        return []
    rules = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        seg = row[idx["segment"]] if idx["segment"] < len(row) else None
        if seg is None:
            continue
        seg = str(seg).strip()
        r_min = _to_int(row[idx["r_min"]] if idx["r_min"] < len(row) else None)
        r_max = _to_int(row[idx["r_max"]] if idx["r_max"] < len(row) else None)
        f_min = _to_int(row[idx["f_min"]] if idx["f_min"] < len(row) else None)
        f_max = _to_int(row[idx["f_max"]] if idx["f_max"] < len(row) else None)
        m_min = _to_int(row[idx["m_min"]] if idx["m_min"] < len(row) else None)
        m_max = _to_int(row[idx["m_max"]] if idx["m_max"] < len(row) else None)
        if None in (r_min, r_max, f_min, f_max, m_min, m_max):
            continue
        rules.append((seg, r_min, r_max, f_min, f_max, m_min, m_max))
    wb.close()
    return rules


def _assign_segment(df: pd.DataFrame, rules: list) -> pd.DataFrame:
    """ستون segment را بر اساس قواعد به df اضافه می‌کند."""
    def find_segment(row):
        r, f, m = row.get("r_score"), row.get("f_score"), row.get("m_score")
        if pd.isna(r) or pd.isna(f) or pd.isna(m):
            return "Unclassified"
        ri, fi, mi = int(r), int(f), int(m)
        for seg, r_min, r_max, f_min, f_max, m_min, m_max in rules:
            if r_min <= ri <= r_max and f_min <= fi <= f_max and m_min <= mi <= m_max:
                return seg
        return "Unclassified"

    df = df.copy()
    df["segment"] = df.apply(find_segment, axis=1)
    return df


def build_rfm_charts(folder: Path) -> tuple[bool, str, list[str]]:
    """
    ساخت پوشه charts و ۷ نمودار.
    برمی‌گرداند: (success, message, list of chart filenames).
    """
    folder = Path(folder)
    scores_file = folder / "rfm_scores.xlsx"
    constant_file = folder / "rfm_constant.xlsx"
    if not scores_file.is_file():
        return False, "فایل rfm_scores.xlsx یافت نشد.", []
    if not constant_file.is_file():
        return False, "فایل rfm_constant.xlsx یافت نشد.", []

    try:
        df = pd.read_excel(scores_file, sheet_name=0)
    except Exception as e:
        return False, f"خطا در خواندن rfm_scores.xlsx: {e!s}", []

    required = {"r_score", "f_score", "m_score", "total_orders", "total_spent", "recency_days"}
    missing = required - set(df.columns)
    if missing:
        return False, f"ستون‌های لازم در rfm_scores ناقص: {missing}", []

    rules = _load_segment_rules(constant_file)
    if not rules:
        return False, "قواعد سگمنت در rfm_constant.xlsx یافت نشد.", []
    df = _assign_segment(df, rules)

    charts_dir = folder / "charts"
    charts_dir.mkdir(parents=True, exist_ok=True)
    generated: list[str] = []

    try:
        # 1) Heatmap R-F (count per r_score x f_score)
        pivot = df.groupby(["r_score", "f_score"]).size().unstack(fill_value=0)
        fig, ax = plt.subplots(figsize=(8, 6))
        im = ax.imshow(pivot.values, cmap="YlOrRd", aspect="auto")
        ax.set_xticks(range(len(pivot.columns)))
        ax.set_xticklabels(pivot.columns)
        ax.set_yticks(range(len(pivot.index)))
        ax.set_yticklabels(pivot.index)
        ax.set_xlabel("F score")
        ax.set_ylabel("R score")
        ax.set_title("RF Segment Heatmap (count)")
        plt.colorbar(im, ax=ax, label="Count")
        plt.tight_layout()
        p1 = charts_dir / "rf_heatmap_segment_count.png"
        plt.savefig(p1, dpi=120, bbox_inches="tight")
        plt.close()
        generated.append("charts/rf_heatmap_segment_count.png")

        # 2) Bar chart segment size
        seg_counts = df["segment"].value_counts()
        fig, ax = plt.subplots(figsize=(10, 5))
        seg_counts.plot(kind="bar", ax=ax)
        ax.set_xlabel("Segment")
        ax.set_ylabel("Count")
        ax.set_title("Segment Size")
        plt.xticks(rotation=45, ha="right")
        plt.tight_layout()
        p2 = charts_dir / "segment_size_bar.png"
        plt.savefig(p2, dpi=120, bbox_inches="tight")
        plt.close()
        generated.append("charts/segment_size_bar.png")

        # 3) Scatter Frequency vs Monetary, color by Recency
        fig, ax = plt.subplots(figsize=(8, 6))
        sc = ax.scatter(
            df["total_orders"],
            df["total_spent"],
            c=df["recency_days"],
            cmap="viridis_r",
            alpha=0.5,
            s=10,
        )
        ax.set_xlabel("Frequency (total_orders)")
        ax.set_ylabel("Monetary (total_spent)")
        ax.set_title("Frequency vs Monetary (color: Recency days)")
        plt.colorbar(sc, ax=ax, label="Recency days")
        plt.tight_layout()
        p3 = charts_dir / "frequency_vs_monetary_scatter.png"
        plt.savefig(p3, dpi=120, bbox_inches="tight")
        plt.close()
        generated.append("charts/frequency_vs_monetary_scatter.png")

        # 4) Revenue contribution by segment
        rev = df.groupby("segment")["total_spent"].sum()
        fig, ax = plt.subplots(figsize=(10, 5))
        rev.plot(kind="bar", ax=ax)
        ax.set_xlabel("Segment")
        ax.set_ylabel("Total spent")
        ax.set_title("Revenue Contribution by Segment")
        plt.xticks(rotation=45, ha="right")
        plt.tight_layout()
        p4 = charts_dir / "revenue_contribution_by_segment.png"
        plt.savefig(p4, dpi=120, bbox_inches="tight")
        plt.close()
        generated.append("charts/revenue_contribution_by_segment.png")

        # 5) Recency distribution for At Risk
        at_risk = df[df["segment"] == "At Risk"]
        fig, ax = plt.subplots(figsize=(8, 5))
        if len(at_risk) == 0:
            ax.text(0.5, 0.5, "No At Risk customers", ha="center", va="center")
        else:
            at_risk["recency_days"].hist(ax=ax, bins=min(30, max(1, len(at_risk) // 5)))
        ax.set_xlabel("Recency (days)")
        ax.set_ylabel("Count")
        ax.set_title("Recency Distribution (At Risk)")
        plt.tight_layout()
        p5 = charts_dir / "at_risk_recency_distribution.png"
        plt.savefig(p5, dpi=120, bbox_inches="tight")
        plt.close()
        generated.append("charts/at_risk_recency_distribution.png")

        # 6) CLV vs RFM Score
        df["rfm_total"] = df["r_score"] + df["f_score"] + df["m_score"]
        fig, ax = plt.subplots(figsize=(10, 6))
        colors = {"Champions": "#2ecc71", "Loyal": "#27ae60", "Potential Loyalist": "#3498db",
                  "New Customers": "#9b59b6", "Promising": "#1abc9c", "Need Attention": "#f39c12",
                  "About to Sleep": "#e67e22", "At Risk": "#e74c3c", "Hibernating": "#95a5a6",
                  "Lost": "#7f8c8d", "Unclassified": "#bdc3c7"}
        seg_groups = df.groupby("segment")
        for seg_name, grp in seg_groups:
            ax.scatter(
                grp["rfm_total"],
                grp["total_spent"],
                label=seg_name,
                color=colors.get(seg_name, "#999999"),
                alpha=0.6,
                s=15,
            )
        ax.set_xlabel("RFM Score (R + F + M)")
        ax.set_ylabel("CLV (Total Spent)")
        ax.set_title("Customer Lifetime Value vs. RFM Score")
        ax.legend(fontsize=7, loc="upper left", framealpha=0.8, ncol=2)
        plt.tight_layout()
        p6 = charts_dir / "clv_vs_rfm_score.png"
        plt.savefig(p6, dpi=120, bbox_inches="tight")
        plt.close()
        generated.append("charts/clv_vs_rfm_score.png")

        # 7) Treemap – segment share (size = customer count, label includes revenue %)
        seg_summary = df.groupby("segment").agg(
            count=("segment", "size"),
            revenue=("total_spent", "sum"),
        )
        seg_summary = seg_summary.sort_values("count", ascending=False)
        total_rev = seg_summary["revenue"].sum()
        labels = []
        for seg_name, row in seg_summary.iterrows():
            rev_pct = row["revenue"] / total_rev * 100 if total_rev else 0
            labels.append(f"{seg_name}\n{int(row['count'])} customers\n{rev_pct:.1f}% revenue")
        palette = plt.cm.Set3.colors
        seg_colors = [palette[i % len(palette)] for i in range(len(seg_summary))]
        fig, ax = plt.subplots(figsize=(12, 7))
        squarify.plot(
            sizes=seg_summary["count"].tolist(),
            label=labels,
            color=seg_colors,
            alpha=0.85,
            text_kwargs={"fontsize": 8, "weight": "bold"},
            ax=ax,
        )
        ax.set_title("Segment Treemap (size = customers, label = revenue %)", fontsize=13)
        ax.axis("off")
        plt.tight_layout()
        p7 = charts_dir / "segment_treemap.png"
        plt.savefig(p7, dpi=120, bbox_inches="tight")
        plt.close()
        generated.append("charts/segment_treemap.png")

    except Exception as e:
        return False, f"خطا در ساخت نمودارها: {e!s}", generated

    return True, "نمودارها در پوشه charts ایجاد شدند.", generated
