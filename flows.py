"""جریان‌های کاری: وارد کردن داده جدید و استفاده از دادهٔ موجود."""
import shutil
from pathlib import Path

import jdatetime
from bidi.algorithm import get_display
from openpyxl import Workbook, load_workbook

from config import DUMP_DIR, OUTPUT_DIR, SQLITE_DB_PATH, TABLE_GROUPS
from core.customer_purchases import (
    CUSTOMER_PURCHASES_VIEW,
    create_customer_purchases_view,
    get_customer_purchases_row_count,
)
from core.db_manager import SQLiteManager
from core.dump_reader import DumpReader
from core.excel_exporter import ExcelExporter
from core.importer import DumpImporter
from core.rfm_charts import build_rfm_charts, _load_segment_rules
from core.rfm_constants import create_rfm_constant_excel
from core.rfm_data import RFM_DATA_TABLE, create_rfm_data_table, get_rfm_data_row_count
from core.user_full_data import (
    USER_FULL_DATA_TABLE,
    create_user_full_data_table,
    get_user_full_data_row_count,
)
from utils.helpers import create_output_folder, write_output_readme


def rtl(text: str) -> str:
    """تبدیل متن فارسی/عربی برای نمایش درست در کنسول."""
    return get_display(text)


def select_dump_file() -> str | None:
    """نمایش لیست فایل‌ها و انتخاب توسط کاربر."""
    reader = DumpReader(DUMP_DIR)
    files = reader.list_files()

    if not files:
        print(rtl(f"هیچ فایل دامپی در پوشه {DUMP_DIR} یافت نشد."))
        print(rtl("فایل‌های مجاز: .sql, .gz, .sql.gz"))
        return None

    print(rtl("\nفایل‌های دامپ موجود:"))
    print("-" * 50)
    for i, f in enumerate(files):
        comp = rtl(" [فشرده]") if f["compressed"] else ""
        print(f"  {i + 1}. {f['name']} ({f['size_mb']} MB){comp}")
    print("-" * 50)

    while True:
        try:
            choice = input(rtl("شماره فایل را وارد کنید (یا Enter برای اولین فایل، 0 برای خروج):  ")).strip()
            if not choice:
                idx = 0
            else:
                idx = int(choice)
                if idx == 0:
                    return None
                idx -= 1
            path = reader.select_file(idx)
            if path:
                return str(path)
            print(rtl("شماره نامعتبر است."))
        except ValueError:
            print(rtl("لطفاً یک عدد وارد کنید."))
        except (KeyboardInterrupt, EOFError):
            print(rtl("\nلغو شد."))
            return None


def _ask_rfm_base_date() -> str:
    """از کاربر مبنای محاسبه RFM را می‌گیرد: از ابتدا (۰) یا تاریخ شمسی (۱)."""
    print(rtl("\nمبنای محاسبات RFM:"))
    print(rtl("  ۰) از ابتدای تراکنش‌ها"))
    print(rtl("  ۱) می‌خوام تاریخ انتخاب کنم"))
    while True:
        try:
            choice = input(rtl("\nانتخاب:  ")).strip()
            if choice == "0":
                return "0"
            if choice == "1":
                date_val = input(rtl("تاریخ شمسی (مثال 1404/01/20):  ")).strip()
                return date_val if date_val else "0"
            print(rtl("لطفاً ۰ یا ۱ وارد کنید."))
        except (KeyboardInterrupt, EOFError):
            return "0"


def run_import_new_data() -> None:
    """وارد کردن دامپ جدید، ساخت viewها، خروجی Excel و کپی دیتابیس به پوشه خروجی."""
    rfm_from_shamsi_date = _ask_rfm_base_date()
    if str(rfm_from_shamsi_date).strip() and str(rfm_from_shamsi_date).strip() != "0":
        print(rtl(f"مبنای محاسبات RFM (شمسی): {rfm_from_shamsi_date}"))
    else:
        print(rtl("مبنای محاسبات RFM (شمسی): از ابتدا"))

    # ۱. خالی کردن دیتابیس موقت
    with SQLiteManager(SQLITE_DB_PATH) as db:
        dropped = db.clear_all_tables()
    print(rtl(f"\nدیتابیس موقت خالی شد ({dropped} جدول حذف شد)."))

    dump_path = select_dump_file()
    if not dump_path:
        return

    reader = DumpReader()
    info = reader.get_info(dump_path)
    print(rtl(f"\nفایل انتخاب شده: {info['name']}"))
    print(rtl(f"حجم: {info['size_mb']} MB"))
    print(rtl(f"فشرده: {'بله' if info['compressed'] else 'خیر'}"))

    prefix = reader.detect_prefix(dump_path)
    if prefix:
        print(rtl(f"پیشوند تشخیص داده شده: '{prefix}'"))
    else:
        print(rtl("پیشوندی تشخیص داده نشد."))

    complete_groups = reader.get_complete_groups(dump_path, prefix) if TABLE_GROUPS else []
    if TABLE_GROUPS:
        print(rtl("\nبررسی لیست‌ها:"))
        for group_name in TABLE_GROUPS:
            status = "detect" if group_name in complete_groups else "not found"
            print(rtl(f"{group_name}: {status}"))

    if complete_groups:
        print(rtl("\nدر حال وارد کردن جداول به دیتابیس موقت..."))
        importer = DumpImporter(SQLITE_DB_PATH)
        result = importer.import_complete_groups(dump_path, complete_groups, prefix)
        print(rtl(f"  جداول ایجاد شده: {result['tables_created']}"))
        print(rtl(f"  دستورات INSERT اجرا شده: {result['inserts_count']}"))
        if result["errors"]:
            print(rtl("  خطاها:"))
            for err in result["errors"][:5]:
                print(rtl(f"    - {err}"))
            if len(result["errors"]) > 5:
                print(rtl(f"    ... و {len(result['errors']) - 5} خطای دیگر"))

    table_row_counts: dict[str, int] = {}
    if complete_groups:
        with SQLiteManager(SQLITE_DB_PATH) as db:
            idx_result = db.ensure_recommended_indexes()
            if idx_result["created"] > 0:
                print(rtl(f"  ایندکس‌های پیشنهادی ایجاد شد ({idx_result['created']} مورد)."))
            table_row_counts = db.get_table_row_counts()

            if "wp" in complete_groups:
                if create_customer_purchases_view(db):
                    count = get_customer_purchases_row_count(db)
                    table_row_counts[CUSTOMER_PURCHASES_VIEW] = count
                    print(rtl(f"  جدول اطلاعات خرید مشتری ایجاد شد ({count} رکورد)."))
                else:
                    print(rtl("  خطا در ایجاد جدول اطلاعات خرید مشتری."))

                if create_user_full_data_table(db):
                    count = get_user_full_data_row_count(db)
                    table_row_counts[USER_FULL_DATA_TABLE] = count
                    print(rtl(f"  جدول user_full_data ایجاد شد ({count} رکورد)."))
                else:
                    print(rtl("  خطا در ایجاد جدول user_full_data."))

                if create_rfm_data_table(db, from_shamsi_date=rfm_from_shamsi_date):
                    count = get_rfm_data_row_count(db)
                    table_row_counts[RFM_DATA_TABLE] = count
                    if str(rfm_from_shamsi_date).strip() and str(rfm_from_shamsi_date).strip() != "0":
                        print(rtl(f"  جدول rfm_data ایجاد شد ({count} رکورد) - از تاریخ شمسی {rfm_from_shamsi_date}."))
                    else:
                        print(rtl(f"  جدول rfm_data ایجاد شد ({count} رکورد) - بدون فیلتر تاریخ."))
                else:
                    print(rtl("  خطا در ایجاد جدول rfm_data."))

    folder_name = prefix.rstrip("_") if prefix else "output"
    output_folder = create_output_folder(OUTPUT_DIR, folder_name)
    generated_files: list[str] = []

    if CUSTOMER_PURCHASES_VIEW in table_row_counts:
        headers = [
            "شناسه کاربر",
            "نام کاربر",
            "ایمیل",
            "شماره موبایل",
            "شناسه سفارش",
            "تاریخ خرید",
            "مبلغ خرید",
            "وضعیت سفارش",
        ]
        with SQLiteManager(SQLITE_DB_PATH) as db:
            exporter = ExcelExporter(db, output_folder)
            paths = exporter.export_view_chunked(
                CUSTOMER_PURCHASES_VIEW,
                output_base_name="user_orders",
                column_headers=headers,
            )
            for p in paths:
                print(rtl(f"فایل Excel: {p.name}"))
                generated_files.append(p.name)

    if USER_FULL_DATA_TABLE in table_row_counts:
        with SQLiteManager(SQLITE_DB_PATH) as db:
            exporter = ExcelExporter(db, output_folder)
            paths = exporter.export_view_chunked(
                USER_FULL_DATA_TABLE,
                output_base_name="user_full_data",
            )
            for p in paths:
                print(rtl(f"فایل Excel: {p.name}"))
                generated_files.append(p.name)

    if RFM_DATA_TABLE in table_row_counts:
        with SQLiteManager(SQLITE_DB_PATH) as db:
            exporter = ExcelExporter(db, output_folder)
            paths = exporter.export_view_chunked(
                RFM_DATA_TABLE,
                output_base_name="rfm_data",
                column_formats={
                    "total_spent": "#,##0",
                    "last_order_amount": "#,##0",
                },
            )
            for p in paths:
                print(rtl(f"فایل Excel: {p.name}"))
                generated_files.append(p.name)

            # ساخت فایل ثابت‌ها/لیبل‌های پیشنهادی RFM برای استفاده کاربر و مراحل بعد
            const_path = create_rfm_constant_excel(db, output_folder)
            print(rtl(f"فایل Excel: {const_path.name}"))
            generated_files.append(const_path.name)

    # کپی دیتابیس موقت به پوشه خروجی
    dest_db = output_folder / "converted.db"
    shutil.copy2(SQLITE_DB_PATH, dest_db)
    print(rtl(f"کپی دیتابیس به پوشه خروجی: {dest_db.name}"))

    write_output_readme(
        output_folder,
        info["name"],
        info["size_mb"],
        complete_groups=complete_groups,
        table_groups=TABLE_GROUPS,
        table_row_counts=table_row_counts,
        rfm_from_shamsi_date=rfm_from_shamsi_date,
        excel_files=generated_files,
    )
    print(rtl(f"\nپوشه خروجی: {output_folder}"))
    print(rtl("فایل README.txt ایجاد شد."))


# ستون‌های لازم در فایل ۱_rfm_data.xlsx
RFM_DATA_REQUIRED_COLUMNS = {
    "user_id",
    "last_order_date",
    "last_order_date_shamsi",
    "total_orders",
    "total_spent",
    "last_order_amount",
    "recency_days",
}

# شیت‌ها و ستون‌های لازم در rfm_constant.xlsx
RFM_CONSTANT_REQUIRED_SHEETS = {"meta", "thresholds"}
RFM_CONSTANT_THRESHOLDS_COLUMNS = {"metric", "score", "min_value", "max_value"}
RFM_SCORE_REQUIRED_METRICS = {"recency_days", "total_orders", "total_spent"}


def _to_float(value):
    if value is None:
        return None
    txt = str(value).strip()
    if not txt:
        return None
    try:
        return float(txt.replace(",", ""))
    except Exception:
        return None


def _excel_sort_key(path: Path):
    """
    مرتب‌سازی فایل‌های chunked مثل:
    1_rfm_data.xlsx, 2_rfm_data.xlsx, ...
    """
    name = path.name
    parts = name.split("_", 1)
    if len(parts) == 2 and parts[0].isdigit():
        return int(parts[0]), name
    return 10**9, name


def _load_rfm_thresholds(constant_file: Path) -> tuple[dict[str, list[tuple[float, float, int]]], str | None]:
    """
    thresholds را از rfm_constant.xlsx می‌خواند.
    خروجی:
      ({metric: [(min, max, score), ...]}, error_message)
    """
    try:
        wb = load_workbook(constant_file, read_only=True, data_only=True)
        if "thresholds" not in wb.sheetnames:
            wb.close()
            return {}, rtl("شیت thresholds در rfm_constant.xlsx یافت نشد.")
        ws = wb["thresholds"]
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            wb.close()
            return {}, rtl("هدر شیت thresholds خالی است.")
        headers = [str(c).strip() if c is not None else "" for c in header_row]
        idx = {h: i for i, h in enumerate(headers)}
        required = ["metric", "min_value", "max_value", "score"]
        if not all(k in idx for k in required):
            wb.close()
            return {}, rtl("ستون‌های لازم برای thresholds کامل نیستند.")

        rules: dict[str, list[tuple[float, float, int]]] = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            metric = row[idx["metric"]] if idx["metric"] < len(row) else None
            if metric is None:
                continue
            metric = str(metric).strip()
            min_v = _to_float(row[idx["min_value"]] if idx["min_value"] < len(row) else None)
            max_v = _to_float(row[idx["max_value"]] if idx["max_value"] < len(row) else None)
            score_v = row[idx["score"]] if idx["score"] < len(row) else None
            if min_v is None or max_v is None or score_v is None:
                continue
            try:
                score_i = int(float(str(score_v)))
            except Exception:
                continue
            rules.setdefault(metric, []).append((min_v, max_v, score_i))
        wb.close()

        for metric in RFM_SCORE_REQUIRED_METRICS:
            if metric not in rules or not rules[metric]:
                return {}, rtl(f"برای معیار {metric} در thresholds قانونی پیدا نشد.")
            rules[metric].sort(key=lambda x: (x[0], x[1]))

        return rules, None
    except Exception as e:
        return {}, rtl(f"خطا در خواندن rfm_constant.xlsx: {e!s}")


def _score_by_rules(value, rules: list[tuple[float, float, int]]) -> int:
    """
    امتیازدهی مقدار بر اساس بازه‌های (min,max,score).
    """
    v = _to_float(value)
    if v is None:
        return 0

    # تطبیق مستقیم بازه
    for min_v, max_v, score in rules:
        if min_v <= v <= max_v:
            return score

    # fallback برای خطاهای گردکردن/مرزها
    if v < rules[0][0]:
        return rules[0][2]
    if v > rules[-1][1]:
        return rules[-1][2]

    # نزدیک‌ترین بازه
    best_score = rules[0][2]
    best_dist = float("inf")
    for min_v, max_v, score in rules:
        center = (min_v + max_v) / 2
        d = abs(v - center)
        if d < best_dist:
            best_dist = d
            best_score = score
    return best_score


def _build_rfm_scores_file(folder: Path) -> tuple[bool, str]:
    """
    از روی فایل‌های rfm_data و rfm_constant امتیاز R/F/M را محاسبه می‌کند
    و فایل rfm_scores.xlsx را می‌سازد.
    """
    folder = Path(folder)
    constant_file = folder / "rfm_constant.xlsx"
    if not constant_file.is_file():
        return False, rtl("فایل rfm_constant.xlsx موجود نیست.")

    rfm_data_files = sorted(folder.glob("*_rfm_data.xlsx"), key=_excel_sort_key)
    if not rfm_data_files:
        single = folder / "rfm_data.xlsx"
        if single.is_file():
            rfm_data_files = [single]
    if not rfm_data_files:
        return False, rtl("فایل‌های rfm_data پیدا نشدند.")

    rules, err = _load_rfm_thresholds(constant_file)
    if err:
        return False, err

    segment_rules = _load_segment_rules(constant_file)

    output_file = folder / "rfm_scores.xlsx"
    wb_out = Workbook(write_only=True)
    ws_out = wb_out.create_sheet("rfm_scores")
    ws_out.append(
        [
            "user_id",
            "r_score",
            "f_score",
            "m_score",
            "rfm_score",
            "segment",
            "recency_days",
            "total_orders",
            "total_spent",
            "last_order_amount",
            "last_order_date",
            "last_order_date_shamsi",
        ]
    )

    required_headers = {
        "user_id",
        "last_order_date",
        "last_order_date_shamsi",
        "total_orders",
        "total_spent",
        "last_order_amount",
        "recency_days",
    }

    try:
        for file_path in rfm_data_files:
            wb_in = load_workbook(file_path, read_only=True, data_only=True)
            if not wb_in.sheetnames:
                wb_in.close()
                continue
            ws = wb_in[wb_in.sheetnames[0]]
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if not header_row:
                wb_in.close()
                continue
            headers = [str(c).strip() if c is not None else "" for c in header_row]
            idx = {h: i for i, h in enumerate(headers)}
            missing = required_headers - set(idx.keys())
            if missing:
                wb_in.close()
                return False, rtl(f"ستون‌های لازم در {file_path.name} ناقص هستند: {missing}")

            for row in ws.iter_rows(min_row=2, values_only=True):
                user_id = row[idx["user_id"]] if idx["user_id"] < len(row) else None
                recency_days = row[idx["recency_days"]] if idx["recency_days"] < len(row) else None
                total_orders = row[idx["total_orders"]] if idx["total_orders"] < len(row) else None
                total_spent = row[idx["total_spent"]] if idx["total_spent"] < len(row) else None
                last_order_amount = row[idx["last_order_amount"]] if idx["last_order_amount"] < len(row) else None
                last_order_date = row[idx["last_order_date"]] if idx["last_order_date"] < len(row) else None
                last_order_date_shamsi = (
                    row[idx["last_order_date_shamsi"]] if idx["last_order_date_shamsi"] < len(row) else None
                )

                r_score = _score_by_rules(recency_days, rules["recency_days"])
                f_score = _score_by_rules(total_orders, rules["total_orders"])
                m_score = _score_by_rules(total_spent, rules["total_spent"])
                rfm_score = f"{r_score}{f_score}{m_score}"

                segment = "Unclassified"
                for seg, sr_min, sr_max, sf_min, sf_max, sm_min, sm_max in segment_rules:
                    if sr_min <= r_score <= sr_max and sf_min <= f_score <= sf_max and sm_min <= m_score <= sm_max:
                        segment = seg
                        break

                ws_out.append(
                    [
                        user_id,
                        r_score,
                        f_score,
                        m_score,
                        rfm_score,
                        segment,
                        recency_days,
                        total_orders,
                        total_spent,
                        last_order_amount,
                        last_order_date,
                        last_order_date_shamsi,
                    ]
                )
            wb_in.close()

        wb_out.save(output_file)
        return True, rtl(f"فایل rfm_scores.xlsx با موفقیت ایجاد شد.")
    except Exception as e:
        return False, rtl(f"خطا در ساخت rfm_scores.xlsx: {e!s}")


def _append_charts_to_readme(folder: Path, chart_files: list[str]) -> None:
    """افزودن لیست نمودارهای تولیدشده به انتهای README.txt در پوشه خروجی."""
    readme_path = Path(folder) / "README.txt"
    if not readme_path.is_file():
        return
    try:
        text = readme_path.read_text(encoding="utf-8")
        suffix = "\n\nنمودارهای تولید شده:\n" + "\n".join(f"  {f}" for f in chart_files) + "\n"
        if "نمودارهای تولید شده" not in text:
            readme_path.write_text(text.rstrip() + suffix, encoding="utf-8")
    except Exception:
        pass


def _validate_rfm_output_folder(folder: Path) -> tuple[bool, str]:
    """
    چک می‌کند که 1_rfm_data.xlsx و rfm_constant.xlsx در پوشه باشند و ستون‌های لازم را داشته باشند.
    برمی‌گرداند: (ok: bool, message: str)
    """
    folder = Path(folder)
    rfm_data_file = folder / "1_rfm_data.xlsx"
    rfm_constant_file = folder / "rfm_constant.xlsx"

    if not rfm_data_file.is_file():
        return False, rtl("فایل 1_rfm_data.xlsx یافت نشد.")
    if not rfm_constant_file.is_file():
        return False, rtl("فایل rfm_constant.xlsx یافت نشد.")

    try:
        wb_data = load_workbook(rfm_data_file, read_only=True, data_only=True)
        if not wb_data.sheetnames:
            wb_data.close()
            return False, rtl("فایل 1_rfm_data.xlsx شیت ندارد.")
        ws = wb_data[wb_data.sheetnames[0]]
        row1 = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        wb_data.close()
        if not row1:
            return False, rtl("فایل 1_rfm_data.xlsx هدر ندارد.")
        headers = {str(c).strip() for c in row1 if c is not None}
        missing = RFM_DATA_REQUIRED_COLUMNS - headers
        if missing:
            return False, rtl(f"فایل 1_rfm_data.xlsx ستون‌های لازم را ندارد: {missing}")
    except Exception as e:
        return False, rtl(f"خطا در خواندن 1_rfm_data.xlsx: {e!s}")

    try:
        wb_const = load_workbook(rfm_constant_file, read_only=True, data_only=True)
        sheet_names = set(wb_const.sheetnames)
        missing_sheets = RFM_CONSTANT_REQUIRED_SHEETS - sheet_names
        if missing_sheets:
            wb_const.close()
            return False, rtl(f"فایل rfm_constant.xlsx شیت‌های لازم را ندارد: {missing_sheets}")
        ws_thr = wb_const["thresholds"]
        row1 = next(ws_thr.iter_rows(min_row=1, max_row=1, values_only=True), None)
        wb_const.close()
        if not row1:
            return False, rtl("شیت thresholds در rfm_constant.xlsx هدر ندارد.")
        headers = {str(c).strip() for c in row1 if c is not None}
        missing = RFM_CONSTANT_THRESHOLDS_COLUMNS - headers
        if missing:
            return False, rtl(f"شیت thresholds ستون‌های لازم را ندارد: {missing}")
    except Exception as e:
        return False, rtl(f"خطا در خواندن rfm_constant.xlsx: {e!s}")

    return True, rtl("فایلها به درستی موجود هستند.")


def run_use_existing_data() -> Path | None:
    """لیست پوشه‌های خروجی، انتخاب یکی توسط کاربر؛ برمی‌گرداند Path پوشه یا None."""
    output_path = Path(OUTPUT_DIR)
    if not output_path.exists():
        output_path.mkdir(parents=True, exist_ok=True)

    subdirs = [p for p in output_path.iterdir() if p.is_dir()]
    # مرتب‌سازی بر اساس تاریخ ایجاد؛ جدیدترین بالا
    subdirs.sort(key=lambda p: p.stat().st_ctime, reverse=True)

    if not subdirs:
        print(rtl(f"هیچ پوشه‌ای در {OUTPUT_DIR} یافت نشد."))
        return None

    print(rtl("\nپوشه‌های خروجی موجود:"))
    print("-" * 50)
    for i, p in enumerate(subdirs):
        try:
            ctime = p.stat().st_ctime
            date_str = jdatetime.datetime.fromtimestamp(ctime).strftime("%Y/%m/%d %H:%M")
        except Exception:
            date_str = "-"
        file_count = sum(1 for x in p.iterdir() if x.is_file())
        print(rtl(f"  {i + 1}. {p.name}  {date_str}  ({file_count})"))
    print("-" * 50)

    while True:
        try:
            choice = input(rtl("شماره پوشه را وارد کنید (۰ برای خروج):  ")).strip()
            if choice == "0":
                return None
            idx = int(choice)
            if 1 <= idx <= len(subdirs):
                chosen = subdirs[idx - 1]
                print(rtl(f"انتخاب شد: {chosen.name}"))
                ok, msg = _validate_rfm_output_folder(chosen)
                if ok:
                    print(msg)
                    score_ok, score_msg = _build_rfm_scores_file(chosen)
                    print(score_msg)
                    if not score_ok:
                        print(rtl("خطا در محاسبه امتیازها؛ فایل rfm_scores ساخته نشد."))
                    else:
                        chart_ok, chart_msg, chart_files = build_rfm_charts(chosen)
                        print(rtl(chart_msg))
                        if not chart_ok:
                            print(rtl("خطا در ساخت نمودارها."))
                        elif chart_files:
                            _append_charts_to_readme(chosen, chart_files)
                else:
                    print(rtl("فایل ها درست نیستند؛ دوباره داده‌ها را وارد کنید."))
                    print(msg)
                return chosen
            print(rtl("شماره نامعتبر است."))
        except ValueError:
            print(rtl("لطفاً یک عدد وارد کنید."))
        except (KeyboardInterrupt, EOFError):
            print(rtl("\nلغو شد."))
            return None
